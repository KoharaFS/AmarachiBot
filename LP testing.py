import os
import threading
import sys
import speech_recognition as sr
import pyttsx3
import nltk
import string
import google.generativeai as genai
import openpyxl
import time
from dotenv import load_dotenv
from datetime import datetime
from nltk.corpus import stopwords
import matplotlib.pyplot as plt
import matplotlib
import traceback

matplotlib.rcParams['font.family'] = 'Segoe UI Emoji'
nltk.download('punkt_tab')

sys.stderr = open(os.devnull, 'w')

# Load environment variables for Gemini API
load_dotenv()
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
print(os.getenv("GOOGLE_API_KEY"))

# Download necessary NLTK data
try:
    nltk.data.find('corpora/stopwords.zip')
    nltk.data.find('tokenizers/punkt.zip')
except LookupError:
    nltk.download('stopwords')
    nltk.download('punkt')

# Initialize Speech Recognizer and Text-to-Speech Engine
recognizer = sr.Recognizer()
tts_engine = pyttsx3.init()  # If in RasPi, put "driverName=eSpeak" inbetween the parentheses

# Load or create Excel file
EXCEL_FILE = "SpeechDataset6.xlsx"  # Changed to the desired file name
try:
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet_all = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet_all = workbook.active
    sheet_all.append([
        "Speech Timestamp", "Correction Timestamp", "Recognized Speech", "Completed Sentence",
        "Total Processing Time (s)", "API Processing Time (s)", "Speech Processing Time (s)",
        "Emotion Timestamp", "Detected Emotion/Sentiment"
    ])

EMOTION_WORD_FILE = "emotion_words.xlsx"
emotion_category_map = {}
try:
    ew_workbook = openpyxl.load_workbook(EMOTION_WORD_FILE)
    ew_sheet = ew_workbook.active
    #print(f"Workbook loaded successfully: {ew_workbook}")
    #print(f"Active sheet: {ew_sheet.title}")
    #print("Initial emotion_category_map:", emotion_category_map)
    header_row = [cell.value for cell in ew_sheet[1]]
    emoji_row = [cell.value for cell in ew_sheet[2]]
    #print(f"Header Row (from list): {header_row}")
    #print(f"Emoji Row (from list): {emoji_row}")
    emotion_emoji_map = {}
    for i in range(0, len(header_row), 2):
        category = header_row[i]
        if i + 1 < len(header_row):
            emoji = header_row[i + 1]
            if category and emoji:
                emotion_emoji_map[category.lower()] = emoji
    #print(f"Emotion Emoji Map: {emotion_emoji_map}")
    for i in range(0, len(header_row), 2):
        category = header_row[i]
        if category and category.lower() in emotion_emoji_map:
            emoji_for_category = emotion_emoji_map[category.lower()]
            for c_idx in range(1, ew_sheet.max_column + 1):
                if ew_sheet.cell(row=1, column=c_idx).value == category:
                    col_idx_words = c_idx
                    for row_idx in range(3, ew_sheet.max_row + 1):
                        word = ew_sheet.cell(row=row_idx, column=col_idx_words).value
                        if word:
                            emotion_category_map[word.lower()] = emoji_for_category
                    break
    #print("Final emotion_category_map:", emotion_category_map)
except FileNotFoundError:
    print(f"Error: {EMOTION_WORD_FILE} not found. Emotion detection from Excel will not work.")
except Exception as e:
    print(f"Error reading {EMOTION_WORD_FILE}: {e}")
    traceback.print_exc()

POSITIVE_WORDS = ["great", "good", "joyful", "wonderful"]
NEGATIVE_WORDS = ["sad", "angry", "bad", "upset"]


# Initialize Gemini Model
model = genai.GenerativeModel("gemini-2.0-flash")


# Function: Speech-to-Text (from Code 1)
def speech_to_text():
    with sr.Microphone() as source:
        print("I'm listening for main input...")
        recognizer.adjust_for_ambient_noise(source, duration=0.5)
        try:
            audio = recognizer.listen(source, timeout=2, phrase_time_limit=8)
            text = recognizer.recognize_google(audio, show_all=False).lower()
            print("You said for main input:", text)
            return text
        except sr.UnknownValueError:
            print("Sorry, could not understand the audio for main input.")
        except sr.RequestError:
            print("Could not request results, check your internet connection for main input.")
        except sr.WaitTimeoutError:
            print("No speech detected for main input, stopping...")
            return ""


# Function: Preprocess Text (from Code 1)
def preprocess_text(text):
    text = text.lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    words = nltk.word_tokenize(text)
    try:
        stop_words = set(stopwords.words('english'))
        filtered_words = [word for word in words if word not in stop_words]
    except LookupError:
        return text
    return ' '.join(filtered_words)


# Function: Autocomplete Text using Gemini API (from Code 1)
def autofill_sentence(broken_sentence):
    if not broken_sentence:
        return "No input detected.", 0.0
    try:
        prompt = (
            f"You are a language assistant trained to reconstruct incomplete sentences while preserving their original meaning. "
            f"Your objective is to aid aphasia patients in their daily conversations by fixing their sentences."
            f"Your task is to correct and fill in the missing words to make the sentence grammatically and semantically correct, "
            f"without changing its intent. One-word sentences must be filled in with its appropriate sentence. Provide only the corrected filled sentence without additional explanations:\n\n"
            f"Broken Sentence: '{broken_sentence}'\n\n"
        )
        gemini_start_time = time.perf_counter()
        response = model.generate_content(prompt)
        gemini_end_time = time.perf_counter()
        gemini_processing_time = round(gemini_end_time - gemini_start_time, 6)
        completed_sentence = response.candidates[0].content.parts[0].text.strip()
        return completed_sentence if completed_sentence else broken_sentence, gemini_processing_time
    except Exception as e:
        print(f"Error with Gemini API: {e}")
        return broken_sentence, 0.0


# Function: Convert Text-to-Speech (from Code 1)
def text_to_speech(text):
    if text:
        try:
            tts_engine.say(text)
            tts_engine.runAndWait()
        except Exception as e:
            print(f"Text-to-Speech error: {e}")


# 😊 Display Emotion Function (from Code 2 - Adjusted for black text on white background)
def display_emotion(text):
    """Displays recognized text and emoji based on emotion from Excel."""
    detected_emotion_category = None
    emoji = "❓"
    theme_color = "black"
    background_color = "white"
    sentiment = "neutral"
    print(f"Displaying emotion for text: '{text}'")

    recognized_words = text.lower().split()  # Split the recognized text into words

    for word, emo in emotion_category_map.items():
        if word in recognized_words:  # Check if the exact emotion word is in the recognized words
            detected_emotion_category = word
            emoji = emo
            print(f"Found emotion word: '{word}', category: '{detected_emotion_category}', emoji: '{emoji}'")
            break  # Stop after the first emotion word is found

    if not detected_emotion_category:
        for word in POSITIVE_WORDS:
            if word in text:
                sentiment = "positive"
                emoji = "🙂"
                theme_color = "green"
                background_color = "lightgreen"
                print(f"Found positive word: '{word}', sentiment: '{sentiment}', emoji: '{emoji}'")
                break
        else:
            for word in NEGATIVE_WORDS:
                if word in text:
                    sentiment = "negative"
                    emoji = "🙁"
                    theme_color = "red"
                    background_color = "lightcoral"
                    print(f"Found negative word: '{word}', sentiment: '{sentiment}', emoji: '{emoji}'")
                    break
            else:
                sentiment = "neutral"
                emoji = "😐"
                theme_color = "black"
                background_color = "white"
                print(f"Sentiment: '{sentiment}', emoji: '{emoji}' (default)")
    fig_emotion = plt.figure(figsize=(6, 6))
    fig_emotion.canvas.manager.set_window_title("Emotion/Sentiment")
    ax = fig_emotion.add_subplot(111, facecolor=background_color)
    ax.text(0.5, 0.9, "EMOTION/SENTIMENT", color="black", fontsize=14, ha="center", fontweight="bold")
    ax.text(0.2, 0.8, "SPEECH", color="black", fontsize=12, ha="center")
    ax.text(0.8, 0.8,
            detected_emotion_category.upper() if detected_emotion_category else sentiment.upper(),
            color="black", fontsize=12, ha="center")
    ax.text(0.5, 0.6, emoji, fontsize=80, ha="center", va="center")
    ax.text(0.5, 0.4, text.upper(), color=theme_color, fontsize=22, ha="center", fontweight="bold")
    ax.text(0.3, 0.2, "RECOGNIZED", color="black", fontsize=12, ha="center")
    ax.text(0.7, 0.2, "DETECTED", color="black", fontsize=12, ha="center")
    ax.text(0.3, 0.1, text.capitalize(), color="black", fontsize=18, ha="center", fontweight="bold")
    ax.text(0.7, 0.1,
            detected_emotion_category.capitalize() if detected_emotion_category else sentiment.capitalize(),
            color="black", fontsize=18, ha="center", fontweight="bold")
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_frame_on(False)
    plt.show(block=False)
    plt.pause(2)  # Reduced pause time
    plt.close(fig_emotion)
    if detected_emotion_category:
        tts_engine.say(f"The detected emotion is related to {detected_emotion_category}")
    elif sentiment == "positive":
        tts_engine.say("The sentiment is positive.")
    elif sentiment == "negative":
        tts_engine.say("The sentiment is negative.")
    else:
        tts_engine.say("The sentiment is neutral.")
    tts_engine.runAndWait()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return timestamp, detected_emotion_category if detected_emotion_category else sentiment #Return the timestamp and the emotion/sentiment

def show_startup_screen():
    fig_power_on = plt.figure(figsize=(3, 3))
    fig_power_on.canvas.manager.set_window_title("Starting...")
    plt.clf()
    ax_power_on = fig_power_on.add_subplot(111, facecolor='black')
    ax_power_on.text(0.5, 0.6, "Powering On...", ha='center', va='center', fontsize=16,
                    fontweight='bold', color='green')
    ax_power_on.text(0.5, 0.3, "🤖", ha='center', va='center', fontsize=40)
    ax_power_on.axis('off')
    plt.tight_layout()
    plt.show(block=False)
    plt.pause(3)
    plt.close(fig_power_on)


def show_shutdown_screen():
    fig_shutdown = plt.figure(figsize=(3, 3))
    fig_shutdown.canvas.manager.set_window_title("Shutting Down...")
    plt.clf()
    ax_shutdown = fig_shutdown.add_subplot(111, facecolor='black')
    ax_shutdown.text(0.5, 0.6, "Shutting Down...", ha='center', va='center', fontsize=16,
                    fontweight='bold', color='red')
    ax_shutdown.text(0.5, 0.3, "🛑", ha='center', va='center', fontsize=40)
    ax_shutdown.axis('off')
    plt.tight_layout()
    plt.show(block=False)
    plt.pause(3)
    plt.close(fig_shutdown)


# Main loop
if __name__ == "__main__":
    show_startup_screen()
    print("Ready to process speech and emotions...")
    fig_listen = None
    update_made = False  # Initialize update_made here

    try:
        while True:
            if fig_listen is None or not plt.fignum_exists(1):
                fig_listen = plt.figure(1, figsize=(3, 3))
                fig_listen.canvas.manager.set_window_title("Listening...")
                plt.clf()
                ax_listen = fig_listen.add_subplot(111, facecolor='black')
                ax_listen.text(0.5, 0.6, "Listening...", ha='center', va='center', fontsize=16,
                                color='white')
                ax_listen.text(0.5, 0.3, "🎤", ha='center', va='center', fontsize=40)
                ax_listen.axis('off')
                plt.tight_layout()
                plt.show(block=False)
                plt.pause(0.001)
            else:
                fig_listen.canvas.manager.set_window_title("Listening...")
                plt.pause(0.001)

            speech_start_time = time.perf_counter()
            spoken_text = speech_to_text()
            speech_end_time = time.perf_counter()
            speech_processing_time = round(speech_end_time - speech_start_time, 6)

            if not spoken_text:
                continue

            if spoken_text.lower() == "shutdown":
                print("Exiting program...")
                if fig_listen is not None and plt.fignum_exists(1):
                    plt.close(fig_listen)
                    fig_listen = None
                show_shutdown_screen()
                break

            speech_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Process speech text with Gemini
            processed_text = preprocess_text(spoken_text)
            corrected_sentence, gemini_processing_time = autofill_sentence(processed_text)
            correction_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            total_processing_time = round(speech_processing_time + gemini_processing_time, 6)

            print("Recognized Speech:", spoken_text)
            print("Corrected Sentence:", corrected_sentence)
            print(f"Gemini Processing Time: {gemini_processing_time} seconds")
            print(f"Speech Processing Time: {speech_processing_time} seconds")
            print(f"Total Processing Time: {total_processing_time} seconds")
            text_to_speech(corrected_sentence)

            # Display emotion based on the original spoken text
            if fig_listen is not None and plt.fignum_exists(1):
                plt.close(fig_listen)
                fig_listen = None
            emotion_timestamp, detected_emotion = display_emotion(spoken_text) # Capture the return values

            # Save data to the Excel file
            sheet_all.append([
                speech_timestamp, correction_timestamp, spoken_text, corrected_sentence,
                total_processing_time, gemini_processing_time, speech_processing_time,
                emotion_timestamp, detected_emotion
            ])
            update_made = True

    except KeyboardInterrupt:
        print("Program interrupted by user.")
    finally:
        if 'workbook' in locals() and 'update_made' in locals() and update_made:
            workbook.save(EXCEL_FILE)
            print(f"Data saved to {EXCEL_FILE}")
