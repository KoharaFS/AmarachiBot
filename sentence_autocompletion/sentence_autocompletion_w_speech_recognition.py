import google.generativeai as genai
import speech_recognition as sr
import os
import openpyxl
import time
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables
load_dotenv()

# Configure Gemini API
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# Initialize Speech Recognizer
recognizer = sr.Recognizer()

# Load or create Excel file
try:
    workbook = openpyxl.load_workbook('SpeechData.xlsx')
    sheet = workbook.active

    # Check if the sheet has the correct headers
    if sheet.cell(row=1, column=5).value != "Total Processing Time (s)":
        sheet.insert_cols(5)
        sheet.cell(row=1, column=5, value="Total Processing Time (s)")

except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Speech Timestamp", "Correction Timestamp", "Recognized Speech", "Completed Sentence", "Total Processing Time (s)", "API Processing Time (s)", "Speech Processing Time (s)"])

# Initialize Gemini Model
model = genai.GenerativeModel("gemini-2.0-flash")

# Autocomplete sentence using Gemini API
def autofill_sentence(broken_sentence):
    if not broken_sentence:
        return "No input detected.", 0.0

    try:
        prompt = f"Fix this sentence that is most likely produced by aphasia patients by filling in the missing words to make it more cohesive. When responding, only provide one corrected sentence without extra text: {broken_sentence}"

        gemini_start_time = time.time()  # Start timer for Gemini processing
        response = model.generate_content(prompt)
        gemini_end_time = time.time()  # End timer after Gemini processes

        gemini_processing_time = round(gemini_end_time - gemini_start_time, 3)  # Gemini processing duration

        completed_sentence = response.candidates[0].content.parts[0].text.strip()
        return completed_sentence if completed_sentence else broken_sentence, gemini_processing_time

    except Exception as e:
        print(f"Error with Gemini API: {e}")
        return broken_sentence, 0.0  # Return original sentence and 0 processing time if API fails

# Main function for speech recognition
def main():
    """Captures speech, processes it, and corrects the sentence."""
    print("Loading...")

    update_made = False  # Track if we need to save Excel updates

    try:
        with sr.Microphone() as source:
            recognizer.adjust_for_ambient_noise(source, duration=1)  # Reduce background noise

            while True:
                try:
                    print("I'm listening...")
                    speech_start_time = time.time()  # Start timer when speech is recognized
                    audio = recognizer.listen(source, timeout=5)  # Listen with timeout
                    print("Processing your words...")

                    # Convert speech to text
                    text = recognizer.recognize_google(audio).lower()
                    speech_end_time = time.time()  # Stop timer after recognition
                    print("You said:", text)

                    # Exit condition
                    if text in ["zero"]:
                        print("Exiting program...")
                        break

                    # Record Speech Timestamp
                    speech_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # Process speech input with Gemini
                    corrected_sentence, gemini_processing_time = autofill_sentence(text)

                    # Record Correction Timestamp (AFTER processing)
                    correction_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # Calculate API processing time
                    speech_processing_time = round(speech_end_time - speech_start_time, 3)

                    # Computes for the overall processing time
                    total_processing_time = round(speech_processing_time + gemini_processing_time, 3)

                    print("Corrected Sentence:", corrected_sentence)
                    print(f"Gemini Processing Time: {gemini_processing_time} seconds")
                    print(f"Speech Processing Time: {speech_processing_time} seconds")
                    print(f"Total Processing Time: {total_processing_time} seconds")

                    # Save to Excel
                    sheet.append([speech_timestamp, correction_timestamp, text, corrected_sentence, total_processing_time, gemini_processing_time, total_processing_time])

                    update_made = True  # Updated only if data was added

                except sr.UnknownValueError:
                    print("Sorry, could not understand the audio.")
                except sr.RequestError:
                    print("Could not request results, check your internet connection.")
                except sr.WaitTimeoutError:
                    print("No speech detected. Please try again.")

    except KeyboardInterrupt:
        print("Program interrupted. Exiting...")

    finally:
        if update_made:
            workbook.save('SpeechData.xlsx')  # Saves only if new data was added
            print("Data saved to SpeechData.xlsx.")

if __name__ == "__main__":
    main()
